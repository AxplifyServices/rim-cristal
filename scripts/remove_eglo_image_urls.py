from pathlib import Path
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
IMPORT_DIR = ROOT / "database" / "import"

EXCEL_FILES = [
    IMPORT_DIR / "monluminaire_ampoule.xlsx",
    IMPORT_DIR / "monluminaire_eclairage_exterieur.xlsx",
    IMPORT_DIR / "monluminaire_eclairage_interieur.xlsx",
]

IMAGE_COLUMNS = [
    "url_image1",
    "url_image2",
    "url_image3",
    "url_image4",
    "url_image5",
]

BAD_PATTERNS = [
    "eglo.png",
    "/img/m/",
    "/img/p/",
    "fr-default",
    "default-large_default",
    "default-home_default",
]


def is_bad_image_url(value):
    if value is None:
        return False

    text = str(value).strip().lower()

    if not text:
        return False

    return any(pattern.lower() in text for pattern in BAD_PATTERNS)


def is_valid_product_image_url(value):
    if value is None:
        return False

    text = str(value).strip()
    lowered = text.lower()

    if not lowered.startswith("http"):
        return False

    if is_bad_image_url(text):
        return False

    if not lowered.endswith((".jpg", ".jpeg", ".png", ".webp")):
        return False

    if "-large_default/" not in lowered and "-home_default/" not in lowered:
        return False

    return True


def clean_image_values(values):
    cleaned = []

    for value in values:
        if not is_valid_product_image_url(value):
            continue

        text = str(value).strip()

        if text not in cleaned:
            cleaned.append(text)

    while len(cleaned) < len(IMAGE_COLUMNS):
        cleaned.append(None)

    return cleaned[:len(IMAGE_COLUMNS)]


def clean_excel_file(path: Path):
    if not path.exists():
        print(f"[MANQUANT] {path}")
        return

    wb = load_workbook(path)

    total_removed = 0
    rows_without_image = 0

    for ws in wb.worksheets:
        headers = {}

        for cell in ws[1]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column

        if not all(col in headers for col in IMAGE_COLUMNS):
            continue

        image_col_indexes = [headers[col] for col in IMAGE_COLUMNS]

        for row_idx in range(2, ws.max_row + 1):
            current_values = [
                ws.cell(row=row_idx, column=col_idx).value
                for col_idx in image_col_indexes
            ]

            bad_count = sum(1 for value in current_values if is_bad_image_url(value))
            total_removed += bad_count

            cleaned_values = clean_image_values(current_values)

            if all(value is None for value in cleaned_values):
                rows_without_image += 1

            for col_idx, new_value in zip(image_col_indexes, cleaned_values):
                ws.cell(row=row_idx, column=col_idx).value = new_value

    wb.save(path)

    print(
        f"[OK] {path.name} : "
        f"{total_removed} mauvaise(s) image(s) supprimée(s), "
        f"{rows_without_image} ligne(s) sans image après nettoyage"
    )


def main():
    print("Nettoyage complet des images dans les fichiers Excel...\n")

    for file in EXCEL_FILES:
        clean_excel_file(file)

    print("\nTerminé.")


if __name__ == "__main__":
    main()